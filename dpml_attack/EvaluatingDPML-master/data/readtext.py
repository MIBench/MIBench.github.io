import xlwt #需要的模块
import os
import openpyxl


# """
# 读取数据，查看数据
# """
# f_path=r'.\texas_result\texas100_texas100result.txt'
# with open(f_path,encoding = "UTF-8") as f:
#     data=f.read()
#     print(data)
#
# """
# 将txt数据转换为xls（表格）文件，方便后面做数据分析
# """
#
# # -*- encoding: utf-8 -*-
#
#
# def txt_xls(filename,xlsname):
# #文本转换成xls的函数
# #param filename txt文本文件名称、
# #param xlsname 表示转换后的excel文件名
#     try:
#         f = open(r".\texas_result\texas100_texas100result.txt",encoding = "UTF-8")
#         xls=xlwt.Workbook()
#         #生成excel的方法，声明excel
#         sheet = xls.add_sheet('sheet1',cell_overwrite_ok=True)
#         x = 0
#         while True:
#             #按行循环，读取文本文件
#             line = f.readline()
#             if not line:
#                 break  #如果没有内容，则退出循环
#             for i in range(len(line.split(': '))):
#                 item=line.split(': ')[i]
#                 sheet.write(x,i,item) #x单元格经度，i 单元格纬度
#             x += 1 #excel另起一行
#         f.close()
#         xls.save(xlsname) #保存xls文件
#     except:
#         raise
if __name__ == "__main__" :
    # path1 = "C:\\Users\\24538\\Desktop\\成员推理实验\\EvaluatingDPML-master-tu\\data\\result"
    # path2 = "./texas_result"
    # path3 = "C:\\Users\\24538\\Desktop\\成员推理实验\\calibration_membership-main\\calibration_membership-main\data\\texas_result"
    # path4 = "C:\\Users\\24538\\Desktop\\成员推理实验\\calibration_membership-main\\calibration_membership-main\\data\\purchase_result"
    # path5 = "C:\\Users\\24538\\Desktop\\成员推理实验\\calibration_membership-main\\calibration_membership-main-u\data\\result"
    # xlsname  = "./result.xlsx"
    #
    # workbook = openpyxl.Workbook()
    # sheet1 = workbook.create_sheet('图_实际')  # create a sheet in the last
    # sheet2 = workbook.create_sheet('texas_实际')  # create a sheet in the last
    #
    # sheet3 = workbook.create_sheet('texas_校准')  # create a sheet in the last
    #
    # sheet4 = workbook.create_sheet('purchase_校准')  # create a sheet in the last
    #
    # sheet5 = workbook.create_sheet('图_校准')  # create a sheet in the last
    # paths = [path1,path2,path3,path4,path5]
    # sheets = [sheet1,sheet2,sheet3,sheet4,sheet5]
    # index = 0
    # for path in paths:
    #     sheet = sheets[index]
    #     files = os.listdir(path)  # 得到文件夹下的所有文件名称
    #     x = 1
    #     for file in files: #遍历文件夹
    #         print(file)
    #         if not os.path.isdir(file):#判断是否是文件夹，不是文件夹才打开
    #             sheet.cell(x,1,file)
    #             f = open(path+"/"+file,encoding = "utf-8") # 打开文件
    #             line = f.readline()
    #             if not line:
    #                 break  # 如果没有内容，则退出循环
    #             for i in range(len(line.split(': '))):
    #                 item = line.split(': ')[i]
    #                 item = str(item)
    #                 print(item)
    #                 sheet.cell(x, i+2, item)  # x单元格经度，i 单元格纬度
    #             f.close()
    #         x += 1
    #     workbook.save(xlsname)
    #
    #     index += 1
    path = "./purchase_result"
    files = os.listdir(path)
    print(len(files))